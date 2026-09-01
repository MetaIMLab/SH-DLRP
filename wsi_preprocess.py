import os
import openpyxl
import re
from PIL import ImageDraw, Image
import lxml.etree as ET
import json

from histolab.tiler import ScoreTiler, GridTiler
from histolab.masks import TissueMask, BinaryMask
from openslide import OpenSlide
from histolab.slide import Slide
import matplotlib.pyplot as plt
import numpy as np


import pandas as pd

def list_generate(svs_path, xml_path, report_name_path):
    svs_files = sorted(f for f in os.listdir(svs_path) if f.endswith(".svs"))
    xml_files = sorted(f for f in os.listdir(xml_path) if f.endswith(".xml"))

    pre = re.compile(u'[\u4e00-\u9fa5]')

    workbook = openpyxl.load_workbook(report_name_path)
    sheet = workbook.active

    print("Excel cases:", sheet.max_row)
    print("Total svs:", len(svs_files))
    print("Total xml:", len(xml_files))

    process_list = []
    for i, report in enumerate(sheet.iter_rows(values_only=True)):
        for svs_file_name in svs_files:
            res_svs = re.findall(pre, svs_file_name)
            res_report = re.findall(pre, str(report[1]))
            patient_name_svs = ''.join(res_svs)
            patient_name_report = ''.join(res_report)
            if patient_name_svs.startswith(patient_name_report):
                for xml_file_name in xml_files:
                    if svs_file_name.replace('.svs', '') == xml_file_name.replace('.xml', ''):
                        item = {'id': i+1,  'patient_name_svs': patient_name_svs, 'patient_name_report': report[1], 'svs_file_path': os.path.join(svs_path, svs_file_name), 'xml_file_path': os.path.join(xml_path, xml_file_name), 'report': report}
                        process_list.append(item)

    return process_list


def xml_to_region(xml_file):
    """
    parse XML label file and get the points
    :param xml_file: xml file
    :return: region list,region_class
    """

    tree = ET.parse(xml_file)


    region_list = []

    for ann in tree.findall('.//Annotation'):

        for region in ann.findall('./Regions/Region'):
            points = [vertex.attrib for vertex in region.findall('./Vertices/Vertex')]
            if points:
                region_list.append(points)
    return region_list

def region_binary_image(tile, vertex_list, level_downsample, npy_file_path, mask_file_path, save_npy=True, save_mask=True):
    """
    convert the region labeled or not by doctor to binary image
    :param tile: a return image based on the method of Slide class object in 'utils.openslide_utils'
    :param region_list: region list, region point,
                    eg : [[{'X': '27381.168113', 'Y': '37358.653791'}], [{'X': '27381.168113', 'Y': '37358.653791'}]]
    :param region_class : list,keep the value of region.attrib.get('Type') in elements of region list
                    eg : [0,0,0,1,2,3]
    :param level_downsample: slide level down sample
    :param label_correction: label correctting or not
    :return: image painted in region line of numpy array format
    """

    im = Image.new(mode="L", size=tile.size, color=0)
    dr = ImageDraw.Draw(im)

    for region in vertex_list:
        point_list = []
        for point in region:
            X, Y = float(point['X']) / level_downsample, float(point['Y']) / level_downsample
            point_list.append((int(round(X)), int(round(Y))))
        if len(point_list) >= 3:
            dr.polygon(point_list, fill=255)

    filter_matrix = np.array(im).astype(np.uint8)

    if save_mask:
        im.save(mask_file_path)

    if save_npy:
        np.save(npy_file_path, filter_matrix)


    return filter_matrix


def xml2npy(process_list, npy_path, mask_path):
    os.makedirs(npy_path, exist_ok=True)
    os.makedirs(mask_path, exist_ok=True)
    for item in process_list:
        npy_file_path = os.path.join(npy_path, item['svs_file_path'].split('/')[-1].replace('.svs', '.npy'))
        mask_file_path = os.path.join(mask_path, item['svs_file_path'].split('/')[-1].replace('.svs', '.jpg'))

        if not os.path.exists(npy_file_path) or not os.path.exists(mask_file_path):
            print("="*80)
            print(item)
            slide = OpenSlide(item['svs_file_path'])
            xml_file_path = item['xml_file_path']


            level = min(2, len(slide.level_dimensions) - 1)
            tile = slide.get_thumbnail(slide.level_dimensions[level])
            vertex_list = xml_to_region(xml_file_path)
            region_binary_image(
                tile,
                vertex_list,
                slide.level_downsamples[level],
                npy_file_path,
                mask_file_path,
                save_npy=True,
                save_mask=True,
            )

        npy_dict = {'npy_file_path':npy_file_path}
        mask_dict = {'mask_file_path':mask_file_path}
        item.update(npy_dict)
        item.update(mask_dict)

    return process_list




class MyCustomMask(BinaryMask):

    def __init__(self, mask_root):
        super().__init__()
        self.mask_root = mask_root

    def _mask(self, slide):
        thumbnail = slide.thumbnail
        thumbnail_size = thumbnail.size[0] * thumbnail.size[1]
        scaled_image = slide.scaled_image()
        scaled_image_size = scaled_image.size[0] * scaled_image.size[1]
        scaled_slide = thumbnail if thumbnail_size > scaled_image_size else scaled_image

        mask = np.load(self.mask_root)
        # my_mask = apply_mask_image(scaled_slide, mask)
        my_mask = mask

        return my_mask


def tile(process_list, resample_path, info_path, patch_path, seg_path, stitch_path):

    for sample in process_list:

        # if os.path.exists(os.path.join(stitch_path, str(sample['id'])+'.png')):
        #     stitch_file_path = os.path.join(stitch_path, str(sample['id'])+'.png')
        #     print("Exist:", stitch_file_path)
        #     seg_dict = {'stitch_file_path': stitch_file_path}
        #     sample.update(seg_dict)
        #     continue

        print("================================================================")
        print("Processing id:", sample['id'])
        print("Processing patient:", sample['patient_name_report'])
        print(sample)


        slide = Slide(sample['svs_file_path'], processed_path=patch_path)


        print("Information extracting...")

        levels = slide.levels

        level_dimensions = [slide.level_dimensions(level) for level in levels]

        info = {"levels": levels, "level_dimensions": level_dimensions}

        with open(os.path.join(info_path, str(sample['id']) + '.json'), 'w', encoding='utf-8') as json_file:
            json.dump(info, json_file, ensure_ascii=False, indent=4)


        print("Resample extracting...")

        resample_path_image,  resample_path_array= slide._resample()

        # width, height = level3_image.size

        # assert width ==  level_dimensions[2][0]
        # assert height == level_dimensions[2][1]

        resample_path_image.save(os.path.join(resample_path, str(sample['id'])+'.png'))


        print("Mask locating...")

        mask = MyCustomMask(mask_root=sample['npy_file_path'])

        seg_img = slide.locate_mask(mask)

        seg_file_path = os.path.join(seg_path, sample['svs_file_path'].split('/')[-1].replace('.svs', '.png'))

        seg_img.save(seg_file_path)
        seg_img.save(os.path.join(seg_path,str(sample['id'])+'.png'))

        seg_dict = {'seg_file_path': seg_file_path}
        sample.update(seg_dict)


        grid_tiles_extractor = GridTiler(
            # scorer=NucleiScorer(),
            tile_size=(224, 224),
            # n_tiles=100,
            level=0,
            check_tissue=False,
            # tissue_percent=80.0,
            pixel_overlap=0,  # default
            # prefix=sample['wsi_file_path'].split('/')[-1].replace('.svs', ''),  # save tiles in the "scored" subdirectory of slide's processed_path
            prefix=str(sample['id'])+'/',
            suffix=".png"  # default
        )

        # print("Tiles extracting...")
        #
        # grid_tiles_extractor.extract(slide, mask)


        print("Tiles locating...")

        stitch_img = grid_tiles_extractor.locate_tiles(slide=slide, extraction_mask=mask)

        stitch_file_path = os.path.join(stitch_path, sample['svs_file_path'].split('/')[-1].replace('.svs', '.png'))
        stitch_img.save(stitch_file_path)
        stitch_img.save(os.path.join(stitch_path, str(sample['id'])+'.png'))

        seg_dict = {'stitch_file_path': stitch_file_path}
        sample.update(seg_dict)

        print("================================================================")
    return process_list


if __name__ == '__main__':

    center = 'sysucc'

    report_name_path = './dataset/BreCLS-CC/report-name.xlsx'

    svs_path = '../datasets/wsi2/' + center + '-svs/'

    xml_path = '../datasets/wsi2/' + center + '-svs/'

    npy_path = '../datasets/wsi2/' + center + '-npy/'

    resample_path = './dataset/' + center + '-resample/'

    info_path = './dataset/' + center + '-info/'

    mask_path = './dataset/' + center + '-mask/'

    patch_path = './dataset/' + center + '-patch'

    seg_path = './dataset/' + center + '-seg/'
    stitch_path = './dataset/' + center + '-stitch/'

    if not os.path.exists(mask_path):
        os.makedirs(mask_path)
    if not os.path.exists(npy_path):
        os.makedirs(npy_path)
    if not os.path.exists(patch_path):
        os.makedirs(patch_path)
    if not os.path.exists(seg_path):
        os.makedirs(seg_path)
    if not os.path.exists(stitch_path):
        os.makedirs(stitch_path)
    if not os.path.exists(resample_path):
        os.makedirs(resample_path)
    if not os.path.exists(info_path):
        os.makedirs(info_path)



    process_list = list_generate(svs_path, xml_path, report_name_path)

    print("Process list len:",len(process_list))
    for z in process_list:
        print(z)

    process_list = xml2npy(process_list, npy_path, mask_path)

    for z in process_list:
        print(z)
    print(len(process_list))

    process_list = tile(process_list, resample_path, info_path, patch_path, seg_path, stitch_path)

    for z in process_list:
        print(z)
    print(len(process_list))

    df = pd.DataFrame.from_records(process_list)
    df.to_excel("./dataset/sysush_info.xlsx")
