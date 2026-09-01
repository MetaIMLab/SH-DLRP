import SimpleITK as sitk
import os
import matplotlib.pyplot as plt
import numpy as np
from radiomics import featureextractor

import pandas as pd

import openpyxl
import csv

from PIL import Image

import cv2




# bus = sitk.ReadImage(bus_path)
# swv = sitk.ReadImage(swv_path)
# swt = sitk.ReadImage(swt_path)
# nii = sitk.ReadImage(nii_path)

# sitk.Show(img, title="cthead1")

# print(img.GetNumberOfComponentsPerPixel())
# print(img)    #meta-data

# img_array = sitk.GetArrayFromImage(img)

# nii_array = sitk.GetArrayFromImage(nii)
#
# for slice in nii_array:
#
#     # unique_labels = set(sitk.GetArrayFromImage(mask_image).ravel())
#     if np.any(slice):
#
#         # slice = np.expand_dims(slice, 2).repeat(3, axis=2)
#         slice = np.expand_dims(slice, 0)
#
#         nii_slice = slice
#         # nii_slice = slice[np.newaxis, :]

# img_new = sitk.GetImageFromArray(img_array)

# nii_new = sitk.GetImageFromArray(nii_slice)

# img = sitk.Cast(img, sitk.sitkFloat32)
# nii_new = sitk.Cast(nii_new, sitk.sitkFloat32)

# size = img.GetSize()
# origin = img.GetOrigin()
# spacing = img.GetSpacing()
# direction = img.GetDirection()

# print(img.GetSize())
# print(nii_new.GetSize())

# img_1 = sitk.VectorIndexSelectionCast(bus,0)
#
# # sitk.Show(img_1)
#
# nii_new.SetOrigin(img_1.GetOrigin())
# nii_new.SetSpacing(img_1.GetSpacing())
# nii_new.SetDirection(img_1.GetDirection())
#
#
#
#
# filter = sitk.LabelStatisticsImageFilter()
# filter.Execute(img_1, nii_new)


def crop_image_2D(image, bbox, depth):
    """Crop a square 2D ROI from a 3D SimpleITK image.

    ``LabelStatisticsImageFilter.GetBoundingBox`` returns
    ``(x, y, z, size_x, size_y, size_z)``, not min/max coordinates.
    """
    if len(bbox) < 6:
        raise ValueError("Expected a 3D SimpleITK bounding box with six values")

    image_size = tuple(int(value) for value in image.GetSize())
    if len(image_size) != 3:
        raise ValueError("Expected a 3D image, got size {}".format(image_size))

    x, y, z, width, height, _ = (int(value) for value in bbox[:6])
    side = min(max(width, height), image_size[0], image_size[1])
    if side <= 0:
        raise ValueError("Bounding box has no positive area: {}".format(bbox))

    center_x = x + width / 2.0
    center_y = y + height / 2.0
    xmin = int(round(center_x - side / 2.0))
    ymin = int(round(center_y - side / 2.0))
    xmin = max(0, min(xmin, image_size[0] - side))
    ymin = max(0, min(ymin, image_size[1] - side))
    zmin = max(0, min(z, image_size[2] - 1))
    depth = max(1, min(int(depth), image_size[2] - zmin))

    image_crop = sitk.RegionOfInterest(
        image,
        [int(side), int(side), depth],
        [xmin, ymin, zmin],
    )

    return image_crop

def extractor_init():
    # 初始化特征提取器的设置
    settings = {}
    settings['binWidth'] = 25
    settings['sigma'] = [3, 5]
    settings['resampledPixelSpacing'] = [1, 1]
    settings['pixelArrayShift'] = 1000
    settings['normalize'] = True
    settings['normalizeScale'] = 100

    # 实例化特征提取器
    extractor = featureextractor.RadiomicsFeatureExtractor(**settings)

    # 指定使用 LoG 和 Wavelet 滤波器
    extractor.enableImageTypeByName('LoG')
    extractor.enableImageTypeByName('Wavelet')
    # 选择所有特征
    extractor.enableAllFeatures()

    return extractor




# 参数1代表标签值，意思是获取标签值为1的区域所在的bounding box；返回[xmin, xmax, ymin, ymax, zmin, zmax]
# bounding_box = filter.GetBoundingBox(1)
# print(bounding_box)
#
# print(swv.GetSize())
#
# print(nii_slice.shape)
#
# nii_sitk = sitk.GetImageFromArray(nii_slice)
#
#
# print(nii_sitk.GetSize())
# crop_image_2D(swv, bounding_box)
#
# print(swv.GetNumberOfComponentsPerPixel())
# print(nii_sitk.GetNumberOfComponentsPerPixel())




# extractor = extractor_init()
# features = extractor.execute(swv, nii_sitk, label=1)
#
# df = pd.DataFrame([features])


def custom_sort_key(str_value):
    digital_res = ""
    digital_flag = False
    sort_list = []
    rank_value = []
    for c in str_value:
        c_ascii = ord(c)
        if c_ascii <= 57 and c_ascii >= 48:
            digital_flag = True
            digital_res += c
        else:
            if digital_flag:
                digital_res = int(digital_res)
                rank_value.append((1, digital_res))
                digital_res = ""
                digital_flag = False
            if c_ascii <= 47 or (c_ascii >= 58 and c_ascii <= 64) or (
                    c_ascii >= 91 and c_ascii <= 96) or c_ascii >= 123:
                # special char
                rank_value.append((0, c_ascii))
            elif c_ascii >= 97 and c_ascii <= 122:
                rank_value.append((2, c_ascii))
            elif c_ascii >= 65 and c_ascii <= 90:
                rank_value.append((3, c_ascii))

            sort_list.extend(rank_value)
            rank_value = []

    return sort_list

def report_generate(report_load, report_save):
    if not os.path.exists(report_save):
        os.makedirs(report_save)


    # 打开Excel文件
    workbook = openpyxl.load_workbook(report_load)
    sheet = workbook.active

    # 遍历每一行数据
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        assert len(row) == 21
        # 创建一个独立的CSV文件
        csv_filename = os.path.join(report_save, "{}.csv".format(i + 1))
        with open(csv_filename, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            # 将行数据写入CSV文件
            writer.writerow(row)

    workbook.close()

def img_transform(report_load, img_root, bus_save, swv_save, swt_save):
    workbook = openpyxl.load_workbook(report_load)
    sheet = workbook.active

    # 遍历excel每一条数据（对应每一个文件夹）
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        folder_path = os.path.join(img_root, "{}".format(i + 1))
        print(f"正在处理文件夹：{folder_path}")

        for filename in os.listdir(folder_path):
            if filename.endswith('.tar'):
                nii_path = os.path.join(folder_path, filename.replace('.tar', ''))
                nii_file = [file for file in os.listdir(nii_path) if file.endswith('.nii.gz')]
                assert len(nii_file) == 1
                nii_file_path = os.path.join(nii_path, nii_file[0])

                roi_img = sitk.ReadImage(nii_file_path, sitk.sitkUInt8)

                roi_array = sitk.GetArrayFromImage(roi_img)

                if len(roi_array.shape) == 3:
                    for roi_slice in roi_array:
                        if np.any(roi_slice):
                            roi_slice = np.expand_dims(roi_slice, 0)
                            roi_sitk = sitk.GetImageFromArray(roi_slice)

                elif len(roi_array.shape) == 2:
                    roi_array = np.expand_dims(roi_array, 0)
                    roi_sitk = sitk.GetImageFromArray(roi_array)

                else:
                    assert 0




        images = [file for file in os.listdir(folder_path) if file.endswith('.IMA')]
        images = sorted(images, key=custom_sort_key)
        assert len(images) == 3

        for idx, image_path in enumerate(images):
            ima_sitk = sitk.ReadImage(os.path.join(folder_path, image_path))

            img_array = sitk.GetArrayFromImage(ima_sitk)


            # assert len(img_array) == 1

            # for ima_slice in img_array:
            #     img_array = ima_slice



            #
            # roi.SetOrigin(img.GetOrigin())
            # roi.SetSpacing(img.GetSpacing())
            # roi.SetDirection(img.GetDirection())

            filter = sitk.LabelStatisticsImageFilter()
            filter.Execute(roi_sitk, roi_sitk)

            bounding_box = filter.GetBoundingBox(1)

            img_crop = crop_image_2D(ima_sitk, bounding_box, 1)

            # sitk.Show(img_crop)
            #
            # print(img_crop.GetSize())

            np_img_crop = sitk.GetArrayFromImage(img_crop)[0]



            # pil_img = Image.fromarray(np_img_crop)

            if not os.path.exists(bus_save):
                os.makedirs(bus_save)
            if not os.path.exists(swv_save):
                os.makedirs(swv_save)
            if not os.path.exists(swt_save):
                os.makedirs(swt_save)

            if idx == 0:
                save_dir = os.path.join(bus_save, "{}.jpg".format(i + 1))
            elif idx == 1:
                save_dir = os.path.join(swv_save, "{}.jpg".format(i + 1))
            elif idx == 2:
                save_dir = os.path.join(swt_save, "{}.jpg".format(i + 1))

            if np_img_crop.ndim == 3 and np_img_crop.shape[2] == 4:
                np_img_crop = cv2.cvtColor(np_img_crop, cv2.COLOR_RGBA2BGRA)
            elif np_img_crop.ndim == 3 and np_img_crop.shape[2] == 3:
                np_img_crop = cv2.cvtColor(np_img_crop, cv2.COLOR_RGB2BGR)
            cv2.imwrite(save_dir, np_img_crop)

    workbook.close()

            # pil_img.save(save_dir)


def extractor_init():
    # 初始化特征提取器的设置
    # settings = {}
    # settings['binWidth'] = 25
    # settings['sigma'] = [3, 5]
    # settings['resampledPixelSpacing'] = [1, 1]
    # settings['pixelArrayShift'] = 1000
    # settings['normalize'] = True
    # settings['normalizeScale'] = 100
    # settings['force2D'] = True
    #
    # # 实例化特征提取器
    # extractor = featureextractor.RadiomicsFeatureExtractor(**settings)
    #
    # # 指定使用 LoG 和 Wavelet 滤波器
    # extractor.enableImageTypeByName('LoG')
    # extractor.enableImageTypeByName('Wavelet')
    # # 选择所有特征
    # extractor.enableAllFeatures()

    settings = {
        'binWidth': 20,
        'sigma': [1, 2, 3],
        'verbose': True,
        'force2D': True,
    }
    extractor = featureextractor.RadiomicsFeatureExtractor(additionInfo=True, **settings)
    extractor.enableImageTypeByName('LoG')  # 即使设置了LoG，但是没有设置sigma的值不会计算该特征
    extractor.enableImageTypeByName('LBP2D')  # 无需其他参数  93个特征
    extractor.enableImageTypeByName('Wavelet')  # 默认情况下就有LL,LH, HL,HH四种滤波 4 * 93 个特征
    extractor.enableFeatureClassByName('shape2D')  # 必须得设置参数
    # 选择所有特征
    extractor.enableAllFeatures()

    return extractor

def feature_generate(report_load, img_root, bus_fea_save, swv_fea_save, swt_fea_save):
    extractor = extractor_init()

    workbook = openpyxl.load_workbook(report_load)
    sheet = workbook.active

    # 遍历excel每一条数据（对应每一个文件夹）
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        folder_path = os.path.join(img_root, "{}".format(i + 1))
        print(f"正在处理文件夹：{folder_path}")

        for filename in os.listdir(folder_path):
            if filename.endswith('.tar'):
                nii_path = os.path.join(folder_path, filename.replace('.tar', ''))
                nii_file = [file for file in os.listdir(nii_path) if file.endswith('.nii.gz')]
                assert len(nii_file) == 1
                nii_file_path = os.path.join(nii_path, nii_file[0])

                roi_img = sitk.ReadImage(nii_file_path, sitk.sitkUInt8)

                roi_array = sitk.GetArrayFromImage(roi_img)

                if len(roi_array.shape) == 3:
                    for roi_slice in roi_array:
                        if np.any(roi_slice):
                            # roi_slice = np.expand_dims(roi_slice, 0)
                            roi_sitk = sitk.GetImageFromArray(roi_slice)

                elif len(roi_array.shape) == 2:
                    # roi_array = np.expand_dims(roi_array, 0)
                    roi_sitk = sitk.GetImageFromArray(roi_array)

                else:
                    assert 0

        images = [file for file in os.listdir(folder_path) if file.endswith('.IMA')]
        images = sorted(images, key=custom_sort_key)
        assert len(images) == 3

        for idx, image_path in enumerate(images):
            ima_sitk = sitk.ReadImage(os.path.join(folder_path, image_path))

            ima_0 = sitk.VectorIndexSelectionCast(ima_sitk , 0)
            ima_1 = sitk.VectorIndexSelectionCast(ima_sitk , 1)
            ima_2 = sitk.VectorIndexSelectionCast(ima_sitk , 2)

            np_ima_0 = sitk.GetArrayFromImage(ima_0)
            np_ima_1 = sitk.GetArrayFromImage(ima_1)
            np_ima_2 = sitk.GetArrayFromImage(ima_2)

            for slice_0, slice_1, slice_2 in zip(np_ima_0, np_ima_1, np_ima_2):
                ima_slice_0, ima_slice_1, ima_slice_2 = slice_0, slice_1, slice_2

            sitk_ima_0 = sitk.GetImageFromArray(ima_slice_0)
            sitk_ima_1 = sitk.GetImageFromArray(ima_slice_1)
            sitk_ima_2 = sitk.GetImageFromArray(ima_slice_2)

            # print("===========")
            # print(sitk_ima_0.GetSize())
            # print(roi_sitk.GetSize())

            features_0 = extractor.execute(sitk_ima_0, roi_sitk, label=1)
            features_1 = extractor.execute(sitk_ima_1, roi_sitk, label=1)
            features_2 = extractor.execute(sitk_ima_2, roi_sitk, label=1)


            if not os.path.exists(bus_fea_save):
                for d in range(3):
                    os.makedirs(os.path.join(bus_fea_save, "{}".format(d)))
            if not os.path.exists(swv_fea_save):
                for d in range(3):
                    os.makedirs(os.path.join(swv_fea_save, "{}".format(d)))
            if not os.path.exists(swt_fea_save):
                for d in range(3):
                    os.makedirs(os.path.join(swt_fea_save, "{}".format(d)))


            if idx == 0:
                save_dir_0 = os.path.join(bus_fea_save, "0/{}.csv".format(i + 1))
                save_dir_1 = os.path.join(bus_fea_save, "1/{}.csv".format(i + 1))
                save_dir_2 = os.path.join(bus_fea_save, "2/{}.csv".format(i + 1))
            elif idx == 1:
                save_dir_0 = os.path.join(swv_fea_save, "0/{}.csv".format(i + 1))
                save_dir_1 = os.path.join(swv_fea_save, "1/{}.csv".format(i + 1))
                save_dir_2 = os.path.join(swv_fea_save, "2/{}.csv".format(i + 1))
            elif idx == 2:
                save_dir_0 = os.path.join(swt_fea_save, "0/{}.csv".format(i + 1))
                save_dir_1 = os.path.join(swt_fea_save, "1/{}.csv".format(i + 1))
                save_dir_2 = os.path.join(swt_fea_save, "2/{}.csv".format(i + 1))

            df_0 = pd.DataFrame([features_0])
            df_1 = pd.DataFrame([features_1])
            df_2 = pd.DataFrame([features_2])

            df_0.to_csv(save_dir_0, index=False)
            df_1.to_csv(save_dir_1, index=False)
            df_2.to_csv(save_dir_2, index=False)


# def gray_feature_generate(report_load, img_root, bus_fea_save, swv_fea_save, swt_fea_save):
#     extractor = extractor_init()
#
#     workbook = openpyxl.load_workbook(report_load)
#     sheet = workbook.active
#
#     # 遍历excel每一条数据（对应每一个文件夹）
#     for i, row in enumerate(sheet.iter_rows(values_only=True)):
#         folder_path = os.path.join(img_root, "{}".format(i + 1))
#         print(f"正在处理文件夹：{folder_path}")
#
#         for filename in os.listdir(folder_path):
#             if filename.endswith('.tar'):
#                 nii_path = os.path.join(folder_path, filename.replace('.tar', ''))
#                 nii_file = [file for file in os.listdir(nii_path) if file.endswith('.nii.gz')]
#                 assert len(nii_file) == 1
#                 nii_file_path = os.path.join(nii_path, nii_file[0])
#
#                 roi_img = sitk.ReadImage(nii_file_path, sitk.sitkUInt8)
#
#                 roi_array = sitk.GetArrayFromImage(roi_img)
#
#                 if len(roi_array.shape) == 3:
#                     for roi_slice in roi_array:
#                         if np.any(roi_slice):
#                             # roi_slice = np.expand_dims(roi_slice, 0)
#                             roi_sitk = sitk.GetImageFromArray(roi_slice)
#
#                 elif len(roi_array.shape) == 2:
#                     # roi_array = np.expand_dims(roi_array, 0)
#                     roi_sitk = sitk.GetImageFromArray(roi_array)
#
#                 else:
#                     assert 0
#
#         images = [file for file in os.listdir(folder_path) if file.endswith('.IMA')]
#         images = sorted(images, key=custom_sort_key)
#         assert len(images) == 3
#
#         for idx, image_path in enumerate(images):
#             ima_sitk = sitk.ReadImage(os.path.join(folder_path, image_path))
#
#             gray_ima = sitk.RGBToLuminance(ima_sitk)
#
#             np_ima = sitk.GetArrayFromImage(gray_ima)
#
#
#             for slice in np_ima:
#                 ima_slice = slice
#
#             sitk_ima = sitk.GetImageFromArray(ima_slice)
#
#             features_gray = extractor.execute(sitk_ima, roi_sitk, label=1)
#
#             if not os.path.exists(os.path.join(bus_fea_save, "gray")):
#                 os.makedirs(os.path.join(bus_fea_save, "gray"))
#             if not os.path.exists(os.path.join(swt_fea_save, "gray")):
#                 os.makedirs(os.path.join(swt_fea_save, "gray"))
#             if not os.path.exists(os.path.join(swv_fea_save, "gray")):
#                 os.makedirs(os.path.join(swv_fea_save, "gray"))
#
#             if idx == 0:
#                 save_dir = os.path.join(bus_fea_save, "gray/{}.csv".format(i + 1))
#             elif idx == 1:
#                 save_dir = os.path.join(swv_fea_save, "gray/{}.csv".format(i + 1))
#             elif idx == 2:
#                 save_dir = os.path.join(swt_fea_save, "gray/{}.csv".format(i + 1))
#
#             df = pd.DataFrame([features_gray])
#
#             df.to_csv(save_dir, index=False)






if __name__ == '__main__':
    img_root = "../datasets/Breast-US/SYSUCC+SYSUSH/ultrasound/"
    # report_load = "../../Datasets/Breast-US/SYSUCC-SYSUSH/report.xlsx"
    # report_save = "./BreCLS/rep/"

    report_load = "./dataset/BreCLS/report-new.xlsx"
    report_save = "./dataset/BreCLS/rep/"

    bus_save = "./dataset/BreCLS/bus/"
    swv_save = "./dataset/BreCLS/swv/"
    swt_save = "./dataset/BreCLS/swt/"

    bus_fea_save = "./dataset/BreCLS/bus_fea/"
    swv_fea_save = "./dataset/BreCLS/swv_fea/"
    swt_fea_save = "./dataset/BreCLS/swt_fea/"

    # img_transform(report_load, img_root, bus_save, swv_save, swt_save)

    report_generate(report_load, report_save)

    # feature_generate(report_load, img_root, bus_fea_save, swv_fea_save, swt_fea_save)






